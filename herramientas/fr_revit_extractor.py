import os
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def extract_revit_data(archivo_revit=None, ruta_salida=None):
    """
    Extrae datos de Revit desde un archivo JSON y genera un Excel.
    Si archivo_revit es 'test' o 'ejemplo', crea datos de prueba.
    """
    try:
        # MODO PRUEBA: Crear datos de ejemplo
        if archivo_revit in ['test', 'ejemplo', 'prueba']:
            print("🔧 MODO PRUEBA: Creando datos de ejemplo...")
            
            # Crear datos de ejemplo
            datos_ejemplo = {
                "muros": [
                    {"Nombre": "Muro Exterior", "Tipo": "Hormigón", "Longitud": 5.5, "Altura": 3.0, "Volumen": 16.5},
                    {"Nombre": "Muro Interior", "Tipo": "Ladrillo", "Longitud": 3.2, "Altura": 2.5, "Volumen": 8.0},
                    {"Nombre": "Muro Fachada", "Tipo": "Aislante", "Longitud": 4.8, "Altura": 3.0, "Volumen": 14.4}
                ],
                "puertas": [
                    {"Nombre": "Puerta Principal", "Tipo": "Metálica", "Ancho": 0.9, "Alto": 2.1, "Material": "Acero"},
                    {"Nombre": "Puerta Interior", "Tipo": "Madera", "Ancho": 0.8, "Alto": 2.0, "Material": "Pino"},
                    {"Nombre": "Puerta Balcón", "Tipo": "Vidrio", "Ancho": 1.2, "Alto": 2.1, "Material": "Aluminio"}
                ],
                "ventanas": [
                    {"Nombre": "Ventana Salon", "Tipo": "Corredera", "Ancho": 1.5, "Alto": 1.2, "Material": "Aluminio"},
                    {"Nombre": "Ventana Dormitorio", "Tipo": "Oscilante", "Ancho": 1.2, "Alto": 1.0, "Material": "PVC"},
                    {"Nombre": "Ventana Baño", "Tipo": "Fija", "Ancho": 0.6, "Alto": 0.6, "Material": "Aluminio"}
                ],
                "suelos": [
                    {"Nombre": "Suelo Planta Baja", "Tipo": "Cerámico", "Area": 45.5, "Material": "Gres"},
                    {"Nombre": "Suelo Planta Primera", "Tipo": "Madera", "Area": 38.2, "Material": "Roble"}
                ]
            }
            
            data = datos_ejemplo
            print("✅ Datos de ejemplo creados")
        
        # MODO NORMAL: Buscar archivo JSON real
        else:
            # Buscar el JSON en ubicaciones comunes
            json_path = None
            temp_dir = os.path.join(os.getenv('TEMP', os.path.expanduser('~\\AppData\\Local\\Temp')), 'revit_ext')
            json_temp = os.path.join(temp_dir, 'revit_data.json')
            
            # Lista de posibles ubicaciones
            posibles_json = [
                json_temp,
                os.path.join(os.getcwd(), 'revit_data.json'),
                os.path.join(os.path.dirname(archivo_revit) if archivo_revit else '', 'revit_data.json'),
                os.path.join(ruta_salida if ruta_salida else '', 'revit_data.json')
            ]
            
            for path in posibles_json:
                if path and os.path.exists(path):
                    json_path = path
                    break
            
            if not json_path:
                return f'Error: No se encontró revit_data.json. Usa "test" como archivo_revit para crear datos de ejemplo'
            
            print(f"📂 Leyendo datos de: {json_path}")
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        
        # Crear Excel
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Colores para cada categoría
        colores = {
            'muros': '4472C4', 
            'suelos': 'FFC000', 
            'techos': '5B9BD5', 
            'puertas': '9E480E',
            'ventanas': '4CAF50',
            'sanitarios': 'FF6B6B',
            'carpinteria': '9E480E'
        }
        
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        total = 0
        hojas = 0
        categorias_procesadas = []
        
        for cat, elems in data.items():
            if not elems or not isinstance(elems, list):
                continue
            
            categorias_procesadas.append(cat)
            ws = wb.create_sheet(cat[:31])
            
            # Obtener todos los headers únicos
            headers = set()
            for elem in elems:
                if isinstance(elem, dict):
                    headers.update(elem.keys())
            headers = sorted(list(headers))
            
            if not headers:
                continue
            
            # Escribir headers
            for col, h in enumerate(headers, 1):
                cell = ws.cell(1, col, h)
                cell.font = Font(bold=True, color='FFFFFF', size=11)
                cell.fill = PatternFill(
                    start_color=colores.get(cat, '999999'), 
                    end_color=colores.get(cat, '999999'), 
                    fill_type='solid'
                )
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Escribir datos
            for row, elem in enumerate(elems, 2):
                if isinstance(elem, dict):
                    for col, h in enumerate(headers, 1):
                        valor = elem.get(h, '')
                        cell = ws.cell(row, col, valor)
                        cell.border = border
                        
                        # Formato para números
                        if isinstance(valor, (int, float)):
                            cell.number_format = '#,##0.00'
            
            # Ajustar ancho de columnas
            for col in range(1, len(headers)+1):
                col_letter = openpyxl.utils.get_column_letter(col)
                ws.column_dimensions[col_letter].width = 15
            
            ws.freeze_panes = 'A2'
            total += len(elems)
            hojas += 1
        
        if hojas == 0:
            return 'Error: No hay datos para exportar'
        
        # Guardar archivo
        if ruta_salida is None:
            ruta_salida = os.getcwd()
        
        os.makedirs(ruta_salida, exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Nombre del archivo
        if archivo_revit == 'test':
            nombre_excel = f'Mediciones_Ejemplo_{ts}.xlsx'
        else:
            nombre_excel = f'Mediciones_Revit_{ts}.xlsx'
        
        archivo_excel = os.path.join(ruta_salida, nombre_excel)
        wb.save(archivo_excel)
        
        # Mostrar resumen
        print("\n" + "="*50)
        print("✅ EXPORTACIÓN EXITOSA")
        print("="*50)
        print(f"📁 Archivo: {archivo_excel}")
        print(f"📊 Hojas creadas: {hojas}")
        print(f"📋 Elementos totales: {total}")
        print(f"📌 Categorías: {', '.join(categorias_procesadas)}")
        print("="*50 + "\n")
        
        return f'OK: {archivo_excel}'
        
    except Exception as e:
        import traceback
        error_msg = f"Error: {str(e)}\n{traceback.format_exc()}"
        print(error_msg)
        return f'Error: {str(e)}'

# Si se ejecuta directamente
if __name__ == "__main__":
    # Probar con datos de ejemplo
    resultado = extract_revit_data('test', 'output')
    print(f"\nResultado: {resultado}")